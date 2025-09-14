"""
Excel Automation Demo using openpyxl and pandas
"""

# ----------------- openpyxl Example -----------------
from openpyxl import Workbook, load_workbook
import pandas as pd

# Step 1: Create Excel file with openpyxl
def create_excel_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Add headers
    ws.append(["Name", "Salary", "Department"])
    # Add rows
    ws.append(["Rahul", 50000, "IT"])
    ws.append(["Sneha", 60000, "HR"])
    ws.append(["Amit", 55000, "Finance"])
    ws.append(["Priya", 65000, "IT"])

    wb.save("employee.xlsx")
    print("✅ employee.xlsx created successfully!")


# Step 2: Read Excel file with openpyxl
def read_with_openpyxl():
    wb = load_workbook("employee.xlsx")
    ws = wb.active

    print("\n📂 Reading employee.xlsx with openpyxl:")
    for row in ws.iter_rows(values_only=True):
        print(row)


# ----------------- pandas Example -----------------
def read_with_pandas():
    df = pd.read_excel("employee.xlsx")
    print("\n📊 Reading employee.xlsx with pandas:")
    print(df)

    # Average salary
    avg_salary = df["Salary"].mean()
    print("\n💰 Average Salary:", avg_salary)

    # Add Bonus column (10% of Salary)
    df["Bonus"] = df["Salary"] * 0.10

    # Save new file
    df.to_excel("employee_with_bonus.xlsx", index=False)
    print("✅ New file saved as employee_with_bonus.xlsx")

    # Filter employees with Salary > 55,000
    high_salary = df[df["Salary"] > 55000]
    print("\n🔥 Employees with high salary (>55,000):")
    print(high_salary)


# ----------------- Main Execution -----------------
if __name__ == "__main__":
    create_excel_file()
    read_with_openpyxl()
    read_with_pandas()
