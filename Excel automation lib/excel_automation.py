from openpyxl import Workbook

# Create a new Excel file
wb = Workbook()
ws = wb.active

# Add some data
ws['A1'] = "Name"
ws['B1'] = "Marks"
ws.append(["Rahul", 85])
ws.append(["Sneha", 92])
ws.append(["Amit", 78])

# Save file
wb.save("students.xlsx")
print("Excel file created successfully!")



# Example 2: Read Excel File

from openpyxl import load_workbook

# Load existing Excel file
wb = load_workbook("students.xlsx")
ws = wb.active

# Read data from the sheet
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row)



# 🔹 Step 2: pandas (Data Analysis with Excel)

# Example 1: Read Excel File

import pandas as pd

# Read Excel file
df = pd.read_excel("students.xlsx")

print(df)        # Show table
print(df.head()) # Show first 5 rows